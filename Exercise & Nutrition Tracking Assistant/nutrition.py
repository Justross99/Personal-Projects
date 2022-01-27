import schedule
import openpyxl
import os
from datetime import date


class Food:
    def __init__(self, calorie, carb, fat, protein):
        self.calorie = int(calorie)
        self.carb = int(carb)
        self.fat = int(fat)
        self.protein = int(protein)

    def __repr__(self):
        return "Food({!r} Calories, {!r} carbs, {!r} fats, {!r} proteins)".format(self.calorie, self.carb, self.fat, self.protein)

    def __add__(self, other):
        cal = self.calorie + other.calorie
        c = self.carb + other.carb
        f = self.fat + other.fat
        p = self.protein + other.protein
        return Food(cal, c, f, p)

    def __new__(cls, cal, carb, fat, protein):
        cal = foodoftheday['B2'].value
        carb = foodoftheday['C2'].value
        fat = foodoftheday['D2'].value
        protein = foodoftheday['E2'].value
        newfood = Food(cal, carb, fat, protein)
        return super().__new__(cls)


def endofday():
    refreshwb()
    if dailyentries["B2"].value:
        newdailyentry()
        cleardailyfoodlist()
        wb.save('Nutrition.xlsx')
    else:
        print("No entry added \:\(")


def newfood():
    refreshwb()
    if foodoftheday['A2'].value:
        newfoodname = foodoftheday['A2'].value
        newfoodcal = foodofthedayro['B2'].value
        newfoodcarb = foodoftheday['C2'].value
        newfoodfat = foodoftheday['D2'].value
        newfoodprotein = foodoftheday['E2'].value
        addedfood = (newfoodname, newfoodcal, newfoodcarb,
                     newfoodfat, newfoodprotein)
        foodoftheday.append(addedfood)
        for i in a_to_e:
            if i == "B":
                foodoftheday["B2"].value = "=(4*C2)+(4*E2)+(9*D2)"
            else:
                foodoftheday[i + "2"] = None
        wb.save('Nutrition.xlsx')
        print("food added!")
    else:
        print("No new food found")


def newdailyentry():
    refreshwb()
    if dailyentries['B2'].value:
        dailycal = dailyentriesro['B2'].value
        dailycarb = dailyentriesro['C2'].value
        dailyfat = dailyentriesro['D2'].value
        dailyprotein = dailyentriesro['E2'].value
        newentry = (today, dailycal, dailycarb, dailyfat, dailyprotein)
        dailyentries.append(newentry)
    wb.save('Nutrition.xlsx')
    print("Daily entry added!")


def cleardailyfoodlist():
    refreshwb()
    newfood()
    dailyfoodcount()
    y = 3
    for q in range(cellcheckcount):
        for l in a_to_e:
            foodoftheday[l + str(y)].value = None
        y += 1
    print("Cleared " + foods_entered + " foods from Foodoftheday")
    wb.save('Nutrition.xlsx')


def dailyfoodcount():
    global cellcheckcount
    cellcheckcount = -2
    foodoftheday["B2"].value = 1
    dailyfoodcounter = foodoftheday["B"]
    for k in dailyfoodcounter:
        if k.value is None:
            break
        cellcheckcount += 1
    foodoftheday["B2"].value = "=(4*C2)+(4*E2)+(9*D2)"
    global foods_entered
    foods_entered = str(cellcheckcount)
    print("There are " + foods_entered + " foods entered today")


def refreshwb():
    global wb
    global wbreadonly
    wb = openpyxl.load_workbook("Nutrition.xlsx")
    wbreadonly = openpyxl.load_workbook("Nutrition.xlsx", data_only=True)


os.chdir("C:\\Users\\rossj\\AppData\\Local\\Programs\\Python\\Python36\\Nutrition")
wb = openpyxl.load_workbook("Nutrition.xlsx")
wbreadonly = openpyxl.load_workbook("Nutrition.xlsx", data_only=True)
wb.template = False
sheetlist = wb.sheetnames
dailyentries = wb['Daily Entries']
dailyentriesro = wbreadonly['Daily Entries']
foodoftheday = wb['Foodoftheday']
foodofthedayro = wbreadonly['Foodoftheday']
schedule.every().day.at("23:55").do(endofday)
schedule.every(10).minutes.do(newfood)
today = str(date.today())
a_to_e = ["A", "B", "C", "D", "E"]
cellcheckcount = 0
foods_entered = ""

print(sheetlist)
print("Functions are : newfood(), newdailyentry(), cleardailyfoodlist(), dailyfoodcount(), and endofday()")
